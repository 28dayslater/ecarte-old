import logging
import pprint

import pymongo
import re
from bson.objectid import ObjectId
from flask.ext.login import current_user
from flask import current_app as smartcarte
from pymongo.errors import DuplicateKeyError
from slugify import slugify
from werkzeug.security import generate_password_hash, check_password_hash

from ecarte.lib import validate

logger = logging.getLogger(__name__)

class User(object):
    def __init__(self, _id, username, email, password=None, is_admin=False):
        self.id = _id
        self.username = username
        self.email = email
        self.password = password or ''
        self.is_admin = is_admin

    def is_authenticated(self):
        return True

    def is_active(self):
        return True

    def is_anonymous(self):
        return False

    def get_id(self):
        return unicode(self.id)

    @staticmethod
    def load(uid):
        u = smartcarte.mongo.db.user.find_one({'_id': ObjectId(uid)})
        if u:
            return User(u['_id'], u['username'], u['email'], u.get('password', ''), u['is_admin'])

    @staticmethod
    def find(name_or_email, exclude_uid=None):
        """ Find a user with the specified name/email.
            If exclude_uid is specified, it is excluded from the search
        """
        u = smartcarte.mongo.db.user.find_one({
            '_id': {'$ne': ObjectId(exclude_uid)},
            '$or': [{'username': name_or_email}, {'email': name_or_email}]})
        if u:
            return User(u['_id'], u['username'], u['email'], u.get('password', ''), u['is_admin'])

    def set_password(self, pwd):
        self.password = generate_password_hash(pwd)

    def check_password(self, pwd):
        return check_password_hash(self.password, pwd)

    def set_temp_password(self):
        import string
        from os import urandom
        chars = string.letters + string.digits
        tmp_pwd = "".join(chars[ord(c) %len(chars)] for c in urandom(6))
        # TODO: remove the log after finishing sending email to the user!
        logger.info("Setting temporary password for user %s: %s", self.username, tmp_pwd)
        self.set_password(tmp_pwd)
        self.temp_password = True

    def save(self):
        data = {
            'username': self.username,
            'email': self.email,
            'password': self.password,
            'is_admin': self.is_admin}
        if self.id:
            data['_id'] = self.id
        try:
            self.id = smartcarte.mongo.db.user.save(data)
            return self.id
        except DuplicateKeyError:
            # unique indexes on username and email
            return 'E_USER_ALREADY_EXISTS'


class Restaurant(object):
    @staticmethod
    def list_names_and_addresses(uid=None):
        crit = dict()
        if uid:
            crit = {'managed_by': ObjectId(uid)}
        print crit
        return smartcarte.mongo.db.restaurant.find(
            filter=crit,
            projection=('name', 'locations.address1', 'locations.address2', 'locations.city', 'locations.province'),
            sort=(('name', pymongo.ASCENDING),))

    @staticmethod
    def get(rid):
        r = smartcarte.mongo.db.restaurant.find_one({'_id': ObjectId(rid)})
        if r and 'managed_by' in r:
            r['managed_by'] = User.load(r['managed_by'])
        return r

    @staticmethod
    def save(r):
        managed_by = r['managed_by']
        if isinstance(r['managed_by'], User):
            r['managed_by'] = r['managed_by'].id
        smartcarte.mongo.db.restaurant.save(r)
        r['managed_by'] = managed_by
        return r

    @staticmethod
    def save_basic_data(r, form):
        data = {key: val.strip() for key,val in form.iteritems()}
        validations = {
            'name': validate.Required,
            'slug': validate.Slug(optional=True),
        }
        if current_user.is_admin:
            validations.update({
                'contact_name': validate.Required,
                'contact_phone': validate.PhoneNumber,
                'contact_email': validate.Email(optional=True),
                'login_username': validate.Username(minlen=6),
                'login_email': validate.Email
            })
        errors = validate.validate_data(data, validations)

        if current_user.is_admin:
            uid = r['managed_by'].id if 'managed_by' in r else None
            if User.find(data['login_username'], uid):
                errors['login_username'] = 'Username in use. Please choose a different one.'
            if User.find(data['login_email'], uid):
                errors['login_email'] = 'Email in use. Please choose a different one.'

        if errors:
            logger.debug('Validation errors: %s , data: %s', errors, form)
            return errors

        if not data.get('slug', None):
            data['slug'] = slugify(data['name'])
        data['languages']= data['languages'].split(',')
        data['description'] = {}
        data['keywords'] = re.split(' *, *', data['keywords'])
        for lang in data['languages']:
            data['description'][lang] = data['description-'+lang]

        for fld in 'name slug languages description keywords contact_name contact_phone contact_email'.split(' '):
            if fld in data:
                r[fld] = data[fld]

        changed_login, new_login = False, False
        if current_user.is_admin:
            if 'managed_by' in r:
                if r['managed_by'].username != data['login_username'] or r['managed_by'].email != data['login_email']:
                    changed_login = True
                r['managed_by'].username = data['login_username']
                r['managed_by'].email = data['login_email']
                r['managed_by'].save()
            else:
                new_login = True
                u = User(None, data['login_username'], data['login_email'])
                u.set_temp_password()
                u.save()
                r['managed_by'] = u

        print '--- Restaurant data:'
        pp = pprint.PrettyPrinter(indent=2)
        pp.pprint(r)
        Restaurant.save(r)
        if changed_login:
            # TODO: send notificatiopn email to the user
            pass
        elif new_login:
            # TODO: send welcome email to the user
            pass


    @staticmethod
    def save_location(r, form):
        if 'locations' not in r:
            r['locations'] = [{}]
        loc = r['locations'][0]
        data = {key: val.strip() for key,val in form.iteritems()}
        validations = {
            'address1': validate.Required,
            'city': validate.Required,
            'order_phone': validate.PhoneNumber,
            'hrs0' : validate.HoursRange(optional=True),
            'hrs1' : validate.HoursRange(optional=True),
            'hrs2' : validate.HoursRange(optional=True),
            'hrs3' : validate.HoursRange(optional=True),
            'hrs4' : validate.HoursRange(optional=True),
            'hrs5' : validate.HoursRange(optional=True),
            'hrs6' : validate.HoursRange(optional=True),
        }
        errors = validate.validate_data(data, validations)
        if errors:
            logger.debug('Validation errors: %s , data: %s', errors, form)
            return errors
        for fld in 'address1 address2 city province order_phone'.split(' '):
            loc[fld] = data[fld]
        loc['hours'] = []
        for idx in range(0,7):
            key = 'hrs%d'%idx
            if data[key] != '':
                loc['hours'].append(data[key].split(' - '))
            else:
                loc['hours'].append([])
        if data['latlong'] != '':
            ll = re.split(' *, *', data['latlong'])
            try:
                loc['lat']  = float(ll[0])
                loc['long'] =  float(ll[1])
            except:
                errors['latlong'] = 'Invalid coordinates'
        else:
            from geopy import geocoders
            from geopy.geocoders.googlev3 import GeocoderQueryError
            g = geocoders.GoogleV3()
            address = loc['address1'] + ' ' + loc['address2'] + ', ' + loc['city'] + ', ' + loc['province'] + ', Canada'
            try:
                addresses = g.geocode(address, exactly_one=False)
                if len(addresses) != 1:
                    logger.warn('Google returned multiple addresses:')
                    for addr in addresses:
                        logger.warn('%s - (%f , %f)' % (addr[0], addr[1][0], addr[1][1]))
                    logger.warn('Provide geo coordinates manually')
                    errors['latlong'] = 'Unable to geo-locate the address. Please specify the coordinates manually.'
                else:
                    loc['lat'], loc['long'] = addresses[0][1][0], addresses[0][1][1]
            except GeocoderQueryError, e:
                # TODO: display warning about geo-location error, ask for manual lat/long or maybe display a map
                logger.warn("The address doesn't exist according to Google.")
                logger.exception('Exception details: ')
                errors['latlong'] = 'Unable to geo-locate the address. Please specify the coordinates manually.'
            except:
                logger.warn('Error in geo-locating the address.')
                logger.exception('Exception details: ')
        if errors:
            return errors
        else:
            #print '--- loc:'
            #pprint.PrettyPrinter(indent=2).pprint(r['locations'][0])
            Restaurant.save(r)


    # @staticmethod
    # def  _check_cat_index(r, parent_idx, idx):
    #     if parent_idx > -1:
    #         if parent_idx >= len(r['menu']):
    #             return False
    #         if 'subcategories' not in r['menu'][parent_idx] or idx >= len(r['menu'][parent_idx]['subcategories']):
    #             return False
    #     else:
    #         if idx >= len(r['menu']):
    #             return False
    #     return True


    @staticmethod
    def save_category(r, form):
        data = Restaurant._prepare_data(form, r['languages'][0])
        idx = int(data.get('index', -1))
        parent_idx = int(data.get('parent_index', -1))

        validations = {
            'category_title': validate.Required
        }
        errors = validate.validate_data(data, validations)
        if errors:
            return errors
        cat = dict(title={}, description={})
        for lang in r['languages']:
            cat['title'][lang] = data['category_title-'+lang]
            cat['description'][lang] = data['category_description-'+lang]

        if parent_idx > -1:
            if idx > -1:
                try:
                    r['menu'][parent_idx]['subcategories'][idx] = cat
                except IndexError:
                    return {'general': 'Category index out of range.'}
            else:
                if 'subcategories' not in r['menu'][parent_idx]:
                    r['menu'][parent_idx]['subcategories'] = []
                r['menu'][parent_idx]['subcategories'].append(cat)
        else:
            if idx > -1:
                try:
                    r['menu'][idx]['title'] = cat['title']
                    r['menu'][idx]['description'] = cat['description']
                except IndexError:
                    return {'general': 'Category index out of range.'}
            else:
                r['menu'].append(cat)

        #pprint.PrettyPrinter(indent=2).pprint(r['menu'])
        # TODO: check for title uniquness
        Restaurant.save(r)


    @staticmethod
    def delete_category(r, parent_idx, idx):
        try:
            if parent_idx > -1:
                del r['menu'][parent_idx]['subcategories'][idx]
            else:
                del r['menu'][idx]
            Restaurant.save(r)
        except IndexError:
            return {'general': 'Category index out of range.'}


    @staticmethod
    def save_entry(r, form):
        data = Restaurant._prepare_data(form, r['languages'][0])
        catidx = int(data.get('catidx', -1))
        subcatidx = int(data.get('subcatidx', -1))
        idx = int(data.get('index', -1))
        try:
            cat = r['menu'][catidx]
            if subcatidx > -1:
                cat = cat['subcategories'][subcatidx]
            if 'entries' not in cat:
                cat['entries'] = []
            if idx < 0:
                cat['entries'].append({})
                entry = cat['entries'][-1]
            else:
                entry = cat['entries'][idx]
            validations = {
                'entry_title': validate.Required,
                'peppers': validate.Value((0, 3))
            }
            errors = validate.validate_data(data, validations)
            ok, pricedata = Restaurant._collect_prices(data)
            if not ok:
                errors.update(pricedata)
            if errors:
                return errors
            if 'title' not in entry:
                entry['title'] = {}
                entry['description'] = {}
            entry['prices'] = []
            for price in pricedata:
                entry['prices'].append({
                    'amount': price['amount'],
                    'description': {}
                })
            for lang in r['languages']:
                entry['title'][lang] = data['entry_title-'+lang]
                entry['description'][lang] = data['entry_description-'+lang]
                for idx in range(0, len(pricedata)):
                    key = 'price_description%d-%s' % (idx+1, lang)
                    entry['prices'][idx]['description'][lang] = data.get(key, '')
            entry['peppers'] = data['peppers']
            markers = data.get('markers','')
            if markers:
                markers_set = set(markers.split(','))
                valid_markers = {'veg', 'gf', 'sf', 'lc'}
                for m in markers_set:
                    if m not in valid_markers:
                        return {'markers': 'Invalid marker: '+m}
            entry['markers'] = markers
            entry['photo'] = data.get('entry_photo', '/img/dish.svg')
#            pprint.PrettyPrinter(indent=2).pprint(cat['entries'])
            Restaurant.save(r)
        except (IndexError, ValueError):
            return {'general': 'Invalid of out of range Category/Entry indexes.'}


    @staticmethod
    def update_entry_image(r, form, doupdate=True):
        try:
            data = (str(r['_id']), form['catidx'], form.get('subcatidx', ''), form['index'], 'jpg')
            filenm = '.'.join(data)
            if doupdate:
                cat = r['menu'][int(data[1])]
                if data[2]:
                    cat = cat['subcategories'][int(data[2])]
                cat['entries'][int(data[3])]['photo'] = filenm
            return True, filenm
        except (IndexError, ValueError) as e:
            #logger.exception('!!!! Fuckup !!!!')
            return False, {'general': 'Invalid or out of range Category/Entry indexes.'}


    _PRICE_REQUIRED = 'Price is required.'

    @staticmethod
    def _collect_prices(form):
        prtab  = [None, None, None, None, None]
        errors = {}
        validator = validate.MonetaryAmount(required=Restaurant._PRICE_REQUIRED, invalid='Invalid price.')
        for idx in range(0,5):
            pk,dk = 'price%d' % (idx+1), 'price_description%d' % (idx+1)
            price, descr = form.get(pk, ''), form.get(dk, '')
            if price or descr:
                ok, data = validator.validate(price)
                if ok:
                    prtab[idx] = { 'amount': data, 'description': descr }
                else:
                    prtab[idx] = errors[pk] = data
        for idx in range(4,-1,-1): # strip empty rows from the end
            if prtab[idx]:
                break
            else:
                del prtab[idx]
        numpr = len(prtab)
        if numpr > 1:
            for idx in range(0,numpr):
                pk = 'price%d' % (idx+1)
                if not prtab[idx]:
                    errors[pk] = 'Price and description are required.'
                elif isinstance(prtab[idx], dict) and not prtab[idx]['description']:
                    errors[pk] = 'Description is required.'
        if errors:
            return False, errors
        else:
            return True, prtab


    @staticmethod
    def _prepare_data(form, defaultlang):
        data = {}
        plpostfix = '-'+defaultlang
        for key,val in form.iteritems():
            data[key] = val.strip()
            if key.endswith(plpostfix): # category_title-en -> category_title
                data[key[:-3]] = form[key]
        return data


########################################################################
# Restaurant data validation and update
# TODO: merge with Restaurant. No need for two classes
########################################################################
class RestaurantAdmin(object):
    class ParseError(Exception):
        def __init__(self, *args):
            self.messages = args

    def __init__(self, rid=None):
        if rid:
            self._r = Restaurant.get(rid)
        else:
            self._r = {}

    @staticmethod
    def update_basic_data(r, data):
        pass

    def import_from_xls(self, filepath):
        import openpyxl
        from openpyxl.cell import cell

        # patch openpyxl's Cell class with a custom get method
        def _val(target):
            ret = target.value
            if not ret:
                return ''
            if isinstance(ret, str) or isinstance(ret, unicode):
                return ret.strip()
            return str(ret)
        setattr(cell.Cell,  'val', _val)

        try:
            workbook = openpyxl.load_workbook(filepath)
        except:
            raise RestaurantAdmin.ParseError("Failed to open data file")
        errors = []
        self._bdata = workbook['Basic data']
        self._menus = workbook['Menu']

        rname = self._bdata.cell('B1').val()
        if rname:
            self._r['name'] = rname
            from slugify import slugify
            self._r['slug'] = slugify(self._r['name'])
        else:
            errors.append('B1: Restaurant name is required')

        strlangs = self._bdata.cell('B17').val()
        if strlangs:
            from babel import Locale, UnknownLocaleError
            langs = re.split(' +| *, *', strlangs) # comma or space separated list
            self._r['languages'] = []
            lset = set()
            for lang in langs:
                try:
                    x = Locale(lang)
                    if lang not in lset: # ignore duplicates
                        self._r['languages'].append(lang)
                        lset.add(lang)
                except UnknownLocaleError:
                    errors.append('B17: invalid language code (%s) (use two letter ISO 639-1 language codes)' % (lang,))
            if len(self._r['languages']) == 0:
                self._r['languages'] = ['en']

        self._r['description'] = { self._r['languages'][0]: self._bdata.cell('B2').val() }

        try:
            self._r['locations'] = [self._parse_address(self._bdata.cell('B3').val())]
        except RestaurantAdmin.ParseError, e:
            errors.extend(e.messages)
            self._r['locations'] = [{}] # continue processing so more errors can be caught
        
        loc = self._r['locations'][0]
        loc['hours'] = []
        for daynum in range(0, 7):
            try:
                text = self._bdata.rows[4][2+daynum].val()
                loc['hours'].append(self._parse_hours(text))
            except RestaurantAdmin.ParseError, e:
                errors.append('%s5: %s' % (chr(ord('C')+daynum), e.messages[0]))

        phtext = self._bdata.cell('B6').val()
        reslt = loc['order_phone'] = self._parse_phone(phtext)
        if not reslt:
            errors.append('B6: Invalid order phone' if phtext else 'B6: 0rder phone is required')

        text = self._bdata.cell('B8').val()
        if text:
            self._r['contact_name'] = ' '.join([w.capitalize() for w in re.split(' +', text)])
        else:
            errors.append('B8: Contact name is required')

        phtext = self._bdata.cell('B9').val()
        if phtext:
            reslt = self._r['contact_phone'] = self._parse_phone(phtext)
            if not reslt:
                errors.append('B9: Invalid contact phone')
        else:
            errors.append('B9: Contact phone is required')

        from validate_email import validate_email
        email = self._bdata.cell('B10').val()
        if email:
            if validate_email(email):
                self._r['contact_email'] = email
            else:
                raise RestaurantAdmin.ParseError('B10: Invalid contact email')

        usrnm = self._bdata.cell('B13').val()
        if usrnm:
            email = self._bdata.cell('B14').val()
            pwd = self._bdata.cell('B15').val()
            login_validation_errors = []
            if not email:
                login_validation_errors.append('B14: Login email is required')
            if not pwd:
                login_validation_errors.append('B15: Pasword is required')
            if login_validation_errors:
                errors.extend(login_validation_errors)
            else:
                if User.find(usrnm):
                    errors.append('B13: username is in use')
                elif User.find(email):
                    errors.append('B14: login email is in use')
                else:
                    u = User(None, usrnm, email, None, False)
                    u.set_password(pwd)
                    self._r['managed_by'] = u

        self._r['menu'], menu_errors = self._parse_menus()
        if menu_errors:
            errors.extend(menu_errors)
        if not errors:
            u = self._r['managed_by']
            self._r['managed_by'] = u.save()
            return smartcarte.db.restaurant.save(self._r), None
        return None, errors

    def _parse_menus(self):
        menu, errors = [], []
        current_section = None
        for rowidx in range(1, len(self._menus.rows[1:])):
            row = self._menus.rows[rowidx]
            if row[0].val():
                menu.append({
                    'title': row[0].val(),
                    'notes': row[3].val()
                })
                current_section = menu[-1]
            elif row[1].val():
                if len(menu) < 1:
                    return ('B%d: Subsection title defined but no sections yet.' % (rowidx+1),)
                if 'subsections' not in menu[-1]:
                    menu[-1]['subsections'] = []
                menu[-1]['subsections'].append({
                    'title': row[1].val(),
                    'notes': row[3].val(),
                })
                current_section = menu[-1]['subsections'][-1]
            elif  row[2].val():
                entry = {
                    'title': row[2].val(),
                    'description': '<br>'.join(row[3].val().split('\n')),
                    'prices': [],
                    'hotness': 0,
                    'vegetarian': False,
                    'vegan': False
                }
                for pr in row[4].val().split('\n'):
                    # number must be either at the start or the end of the text
                    m = re.match('(\d{1,3}(?:\.\d{1,2})?)( *-? *)?', pr)
                    if m:
                        entry['prices'].append((float(m.group(1)), pr[m.end():]))
                    else:
                        m = re.search('( *-? *)?(\d{1,3}(?:\.\d{1,2})?)$', pr)
                        if m:
                            entry['prices'].append((float(m.group(2)), pr[:m.start()]))
                        else:
                            errors.append('E%d: Invalid price data' % (rowidx+1))
                            break
                veg = row[5].val()
                if veg:
                    entry['vegetarian'] = True
                    if veg.lower() == 'vegan':
                        entry['vegan'] = True
                hotness = row[6].val()
                if hotness:
                    if not re.match('^[0-3]', hotness):
                        errors.append('G%d: The level of hotness must 0-3' % (rowidx+1))
                    else:
                        entry['hotness'] = int(hotness)
                if 'entries' not in current_section:
                    current_section['entries'] = []
                current_section['entries'].append(entry)
            else:
                for cell in row[3:]: # allow blank rows
                    if cell.val():
                        errors.append('Invalid data in row %d: expected section/subsection/entry title' % (rowidx+1))
        return (None, errors) if errors else (menu, None)

    def _parse_phone(self, text):
        m = re.match('^\(?([0-9]{3})(?:\) |[ -])?([0-9]{3})-?([0-9]{4})', text)
        if m:
            return '(%s) %s-%s' % m.groups()

    def _parse_address(self, address_line):
        addrparts = address_line.split(',')
        if len(addrparts) not in (3, 4):
            raise RestaurantAdmin.ParseError('B3: invalid address. Expected format: line1(, line2), city, province code')
        loc, errors = {}, []
        addr1 = addrparts[0].strip()
        addr2 = addrparts[1].strip() if len(addrparts) == 4 else ''
        city = addrparts[-2].strip()
        province = addrparts[-1].strip()
        if not addr1:
            errors.append('Address is required')
        if not city:
            errors.append('City is required')
        if not province:
            errors.append('Province code is required.')
        loc['address1'] = ' '.join([w if re.match('^[A-Z]{2}$', w) else w.capitalize() for w in re.split(' +', addr1)])
        loc['address2'] = ' '.join([w if re.match('^[A-Z]{2}$', w) else w.capitalize() for w in re.split(' +', addr2)])
        loc['city'] = city.capitalize()
        loc['province'] = province.upper()
        if loc['province'] not in ('AB', 'BC', 'MB', 'NB', 'NL', 'NS', 'NT', 'NU', 'ON', 'PE', 'QC', 'SK', 'YT'):
            errors.append('Invalid province code')
        if errors:
            raise RestaurantAdmin.ParseError(*errors)
        # Google geocoding is not all that good in Canada.
        # It may return multiple wrong addresses (ex.: 10011 89 ave nw edmonton, ab)
        # TODO: try geocoder.ca
        from geopy import geocoders
        from geopy.geocoders.googlev3 import GQueryError
        g = geocoders.GoogleV3()
        try:
            addresses = g.geocode(address_line, exactly_one=False)
            if len(addresses) != 1:
                print 'Google returned multiple addresses:'
                for addr in addresses:
                    print '%s - (%f , %f)' % (addr[0], addr[1][0], addr[1][1])
                print 'Provide geo coordinates manually'
            else:
                loc['lat'], loc['long'] = addresses[0][1][0], addresses[0][1][1]
        except GQueryError:
            print "The address doesn't exist according to Google. You have to provide lat/long manually."
        except Exception, ex:
            # TODO: log exception
            print 'Google geocoder returned an error: ', ex
        return loc

    def _parse_hours(self, text):
        import time
        if re.match('^ *(closed)? *$', text):
            return 'Closed'
        m = re.match('^ *([0-9]{1,2})(:[0-9]{2})?(am|pm) *- *([0-9]{1,2})(:[0-9]{2})?(am|pm) *',
                     text,
                     re.IGNORECASE)
        if m:
            tmp = list(m.groups())
            if not tmp[1]:
                tmp[1] = ':00'
            if not tmp[4]:
                tmp[4] = ':00'
            strhr1, strhr2 = tmp[0]+tmp[1]+tmp[2].lower(), tmp[3]+tmp[4]+tmp[5].lower()
            hr1, hr2 = None, None
            try:
                hr1 = time.strptime(strhr1, '%I:%M%p')
            except ValueError:
                raise RestaurantAdmin.ParseError('Invalid hour: ' + strhr1)
            try:
                hr2 = time.strptime(strhr2, '%I:%M%p')
            except ValueError:
                raise RestaurantAdmin.ParseError('Invalid hour: ' + strhr2)
            if hr1 >= hr2:
                raise RestaurantAdmin.ParseError('Invalid hour range: ' + text)
            return strhr1, strhr2
        else:
            raise RestaurantAdmin.ParseError('Invalid hours: ' + text)

    # def parse_text_data_and_save(self, lines):
    #     # name - required
    #     m = re.match('^name:\s*(\w+.*)$', lines[0])
    #     if not m:
    #         raise RestaurantAdmin.ParseError('Restaurant name is required.')
    #     self._r['name'] = m.group(1).strip()
    #
    #     # slug - auto-generated if left blank
    #     m = re.match('^slug:\s*(\S+)?', lines[1])
    #     if m:
    #         slugstr = m.group(1).strip() if m.group(1) else None
    #         if slugstr:
    #             m = re.match('^[a-z0-9]+$|^[a-z09]+(-[a-z0-9]+)+$', slugstr)
    #             if not m:
    #                 raise RestaurantAdmin.ParseError('Invalid slug.')
    #             self._r['slug'] = m.group()
    #         else:
    #             from slugify import slugify
    #             self._r['slug'] = slugify(self._r['name'])
    #         lines[1] = 'slug: ' + self._r['slug']
    #     else:
    #         raise RestaurantAdmin.ParseError('Missing slug line.')
    #
    #     # languages - required (default to en)
    #     m = re.match('^languages: *([a-z]{2}( +[a-z]{2})*)$', lines[2])
    #     if m:
    #         # TODO: validate language codes - use babel
    #         self._r['languages'] = m.group(1).split()
    #     else:
    #         raise RestaurantAdmin.ParseError('Missing languages line or invalid language codes.')
    #
    #     # description - optional
    #     m = re.match('^description:\s*(.*)$', lines[3])
    #     if m:
    #         self._r['description'] = {self._r['languages'][0]: m.group(1).strip()}
    #     else:
    #         raise RestaurantAdmin.ParseError('Missing description line.')
    #
    #     #keywords - optional
    #     m = re.match('^keywords: *([a-zA-Z]+( +[a-zA-Z]+)*)? *$', lines[4])
    #     if m:
    #         if m.group(1):
    #             kw = m.group(1).strip().split()
    #             self._r['keywords'] = { self._r['languages'][0]: kw }
    #         else:
    #             self._r['keywords'] = { self._r['languages'][0]: '' }
    #     else:
    #         raise RestaurantAdmin.ParseError('Missing keywords line or invalid keywords.')
    #
    #     # TODO: If multiple locations, 'Add location' button that will add a textarea for the location data
    #     # address -required  TODO: allow address1, address2, address3... - will have to loop
    #     if 'locations' not in self._r:
    #         self._r['locations'] = [{}]
    #     m = re.match('^address: *(\d+.+)$', lines[5])
    #     if m:
    #         self._r['locations'][0].update(self._parse_address(m.group(1).strip(), lines[6]))
    #     else:
    #         raise RestaurantAdmin.ParseError('Address is required.')
    #
    #     # hours - required
    #     m = re.match('^hours: *(\S.*)$', lines[7])
    #     if m and m.group(1):
    #         self._r['locations'][0]['hours'] = RestaurantAdmin._parse_hours(m.group(1))
    #     else:
    #         raise RestaurantAdmin.ParseError('Invalid or missing hours.')
    #
    #     # order phone - required
    #     m = re.match('order phone: *(.+) *', lines[8])
    #     if m:
    #         self._r['locations'][0]['order_phone'] = self._parse_phone(m.group(1))
    #     else:
    #         raise RestaurantAdmin.ParseError('Order phone is required.')
    #
    #     # contact - required
    #     m = re.match('^contact name: *((?:[a-z]+)(?: *[a-z]+)+) *$', lines[9], re.IGNORECASE)
    #     if m:
    #         self._r['locations'][0]['contact_name'] = ' '.join([w.capitalize() for w in re.split(' +', m.group(1))])
    #     else:
    #         raise RestaurantAdmin.ParseError('Missing contact name line.')
    #
    #     # contact phone - optional
    #     m = re.match('^contact phone: *(.+)? *$', lines[10])
    #     if m:
    #         if m.group(1):
    #             self._r['locations'][0]['contact_phone'] = RestaurantAdmin._parse_phone(m.group(1))
    #     else:
    #         raise RestaurantAdmin.ParseError('Missing contact phone line.')
    #     uid = self._r['locations'][0].get('user', None)
    #
    #     # contact/login email - required
    #     m = re.match('^contact email: *(\S+) *$', lines[11])
    #     if m:
    #         email = m.group(1)
    #         from validate_email import validate_email
    #         if validate_email(email):
    #             if User.find(email, uid):
    #                 raise RestaurantAdmin.ParseError('Email already exists.')
    #             self._r['locations'][0]['contact_email'] = email
    #         else:
    #             raise RestaurantAdmin.ParseError('Invalid email')
    #     else:
    #         raise RestaurantAdmin.ParseError('Missing contact email.')
    #
    #     # login
    #     m = re.match('^login username: *([a-zA-Z][a-zA-Z0-9]+) *$', lines[12])
    #     if m:
    #         usrname = m.group(1)
    #         if User.find(usrname, uid):
    #             raise RestaurantAdmin.ParseError('Username already exists.')
    #         if uid:
    #             usr = User.load(uid)
    #             usr.username = usrname
    #             usr.email = email
    #         else:
    #             usr = User(None, usrname, email, '', False)
    #     else:
    #         raise RestaurantAdmin.ParseError('Missing or invalid login username line.')
    #
    #     m = re.match('^password: *(\S+) *$', lines[13])
    #     if m:
    #         usr.set_password(m.group(1))
    #     elif not uid:  # blank password in existing data means no change
    #         raise RestaurantAdmin.ParseError('Password is required.')
    #
    #     self._parse_menus(lines)
    #     usr.save()
    #     self._r['locations'][0]['user'] = usr.id
    #     rid = self.db.restaurant.save(self._r)
    #     return unicode(rid)
